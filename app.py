import os
import json
import pandas as pd
from flask import Flask, render_template, request, jsonify, send_file, session, redirect, url_for
from flask_sqlalchemy import SQLAlchemy
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime
from werkzeug.exceptions import HTTPException
import io
from sqlalchemy import text
import gc
import tempfile
import csv
from openpyxl import load_workbook

def safe_json_dumps(obj):
    def converter(o):
        if hasattr(o, 'isoformat'):
            return o.isoformat()
        if hasattr(o, 'to_dict'):
            return o.to_dict()
        return str(o)
    return json.dumps(obj, default=converter, ensure_ascii=False)

# Configuración de rutas
BASE_DIR = os.path.abspath(os.path.dirname(__file__))
TEMPLATE_DIR = os.path.join(BASE_DIR, "templates")

app = Flask(__name__, template_folder=TEMPLATE_DIR)
# Aseguramos ruta absoluta para SQLite en Render
db_path = os.path.join(BASE_DIR, "anton.db")
app.config['SQLALCHEMY_DATABASE_URI'] = f'sqlite:///{db_path}'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'anton_production_secret_key_2025')
app.config['MAX_CONTENT_LENGTH'] = 32 * 1024 * 1024 # Limit 32MB
db = SQLAlchemy(app)

# --- MODELOS DE DATOS ---

class FiltroMaestro(db.Model):
    __tablename__ = 'filtros_maestros'
    id = db.Column(db.Integer, primary_key=True)
    entidad = db.Column(db.String(50), nullable=False) # 'JEFATURA', 'SUBPROYECTO', 'TECNICO', 'ESTADO'
    valor = db.Column(db.String(100), nullable=False)
    __table_args__ = (db.UniqueConstraint('entidad', 'valor', name='_entidad_valor_uc'),)

class GestionObra(db.Model):
    __tablename__ = 'gestion_obras'
    id = db.Column(db.Integer, primary_key=True)
    data_json = db.Column(db.Text, nullable=False)
    fecha_carga = db.Column(db.DateTime, default=datetime.utcnow)

class ColumnaManual(db.Model):
    __tablename__ = 'columnas_manuales'
    id = db.Column(db.Integer, primary_key=True)
    nombre = db.Column(db.String(50), unique=True, nullable=False)
    tipo = db.Column(db.String(20), nullable=False) # 'fecha', 'lista', 'texto'
    opciones = db.Column(db.Text, nullable=True) # Para tipo 'lista', separado por comas

class ConfiguracionFiltro(db.Model):
    __tablename__ = 'configuracion_filtros'
    id = db.Column(db.Integer, primary_key=True)
    columna = db.Column(db.String(100), unique=True, nullable=False)
    tipo = db.Column(db.String(20), default='search') # 'search' o 'list'
    virtual_cols_json = db.Column(db.Text, nullable=True, default='[]')

class FiltroVirtual(db.Model):
    __tablename__ = 'filtros_virtuales'
    id = db.Column(db.Integer, primary_key=True)
    nombre = db.Column(db.String(50), unique=True, nullable=False) # Ej: "MERCADO", "SEGMENTO"

class MapeoFiltro(db.Model):
    __tablename__ = 'mapeos_filtros'
    id = db.Column(db.Integer, primary_key=True)
    columna_criterio = db.Column(db.String(50), nullable=False) # Ej: "JEFATURA"
    valor_criterio = db.Column(db.String(100), nullable=False) # Ej: "TUMBES"
    valores_json = db.Column(db.Text, nullable=False) # JSON: {"MERCADO": "NORTE", "ZONA": "Z1"}

class Usuario(db.Model):
    __tablename__ = 'usuarios'
    id = db.Column(db.Integer, primary_key=True)
    usuario = db.Column(db.String(50), unique=True, nullable=False)
    nombres = db.Column(db.String(100), nullable=False)
    rol = db.Column(db.String(50), nullable=False) # Ej: "Admin", "Lector"
    activo = db.Column(db.Boolean, default=True)
    password_hash = db.Column(db.String(256), nullable=True)
    restricciones = db.relationship('RestriccionUsuario', backref='usuario_rel', cascade='all, delete-orphan')

class RestriccionUsuario(db.Model):
    __tablename__ = 'restricciones_usuario'
    id = db.Column(db.Integer, primary_key=True)
    usuario_id = db.Column(db.Integer, db.ForeignKey('usuarios.id'), nullable=False)
    columna = db.Column(db.String(100), nullable=False)  # Ej: "JEFATURA"
    valores_json = db.Column(db.Text, nullable=False)    # Ej: '["LIMA", "NORTE"]'

class ConfiguracionGlobal(db.Model):
    __tablename__ = 'configuracion_global'
    clave = db.Column(db.String(50), primary_key=True)
    valor = db.Column(db.Text, nullable=False)

# --- INICIALIZACIÓN ---
try:
    with app.app_context():
        db.create_all()
        
        # Inicializar toggle si no existe
        if not ConfiguracionGlobal.query.get('import_filter_active'):
            db.session.add(ConfiguracionGlobal(clave='import_filter_active', valor='true'))
            
        db.session.commit()
        print("Base de datos inicializada correctamente.")
except Exception as e:
    print(f"Error al inicializar la base de datos: {e}")

# --- GLOBAL ERROR HANDLER ---
@app.errorhandler(Exception)
def handle_global_exception(e):
    if isinstance(e, HTTPException):
        return e
    import traceback
    error_trace = traceback.format_exc()
    print(f"CRASH DETECTADO:\n{error_trace}")
    return jsonify({
        "error": "Error interno del servidor",
        "message": str(e),
        "trace": error_trace
    }), 500

@app.route('/ping')
def ping():
    return jsonify({"status": "ok", "time": datetime.now().isoformat()})

# --- HELPERS DE SESIÓN ---

from functools import wraps

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        user_id = session.get('user_id')
        if not user_id:
            return redirect(url_for('route_login'))
            
        # Verificar que el usuario aún exista en la base de datos
        # Esto previene "Ghost Sessions" en Render tras reinicios de SQLite
        u = db.session.get(Usuario, user_id)
        if not u:
            session.clear()
            return redirect(url_for('route_login'))
            
        return f(*args, **kwargs)
    return decorated

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        user_id = session.get('user_id')
        if not user_id:
            return redirect(url_for('route_login'))
            
        u = db.session.get(Usuario, user_id)
        if not u:
            session.clear()
            return redirect(url_for('route_login'))
            
        rol_upper = (u.rol or "").upper()
        if rol_upper not in ['ADMIN', 'ADMINISTRADOR']:
            return redirect(url_for('route_proyectos'))
            
        # Refrescar sesión para consistencia con frontend
        session['user_id'] = u.id
        session['user_rol'] = u.rol
        return f(*args, **kwargs)
    return decorated

# --- CONFIGURACIÓN DE CONTEXTO ---

@app.context_processor
def inject_current_user():
    """Inyecta current_user en todos los templates."""
    user_id = session.get('user_id')
    if user_id:
        try:
            u = db.session.get(Usuario, user_id)
            if u:
                return {'current_user': {'id': u.id, 'usuario': u.usuario, 'nombres': u.nombres, 'rol': u.rol}}
            else:
                # Si hay ID en sesion pero no en DB, limpiamos para evitar inconsistencias
                session.clear()
        except:
            pass
    return {'current_user': None}

# --- RUTAS DE NAVEGACIÓN ---

@app.route('/login')
def route_login():
    if 'user_id' in session:
        return redirect('/')
    error = request.args.get('error')
    return render_template('login.html', error=error)

@app.route('/logout')
def route_logout():
    session.clear()
    return redirect(url_for('route_login'))

@app.route('/')
@login_required
def route_proyectos():
    return render_template('proyectos.html')

@app.route('/importar')
@admin_required
def route_importar():
    return render_template('importar.html')

@app.route('/tablas')
@admin_required
def route_tablas():
    return render_template('tablas.html')

@app.route('/filtros')
@admin_required
def route_filtros():
    return render_template('filtros.html')

@app.route('/usuarios')
@admin_required
def route_usuarios():
    return render_template('usuarios.html')

@app.route('/dashboard')
@login_required
def route_dashboard():
    return render_template('dashboard.html')

@app.route('/pdt_pago')
@login_required
def route_pdt_pago():
    return render_template('pdt_pago.html')

@app.route('/cerradas')
@login_required
def route_cerradas():
    return render_template('cerradas.html')



# --- API: LOGIN ---

@app.route('/api/login', methods=['POST'])
def api_login():
    data = request.json
    username = (data.get('usuario') or '').strip()
    password = (data.get('password') or '').strip()

    user = Usuario.query.filter_by(usuario=username, activo=True).first()
    if not user:
        return jsonify({'success': False, 'message': 'Usuario no encontrado o inactivo.'}), 401
    if not user.password_hash:
        return jsonify({'success': False, 'message': 'El usuario no tiene contraseña configurada. Contacte al administrador.'}), 401
    if not check_password_hash(user.password_hash, password):
        return jsonify({'success': False, 'message': 'Contraseña incorrecta.'}), 401

    session.permanent = True
    session['user_id'] = user.id
    session['user_rol'] = user.rol
    return jsonify({'success': True, 'redirect': '/'})

# --- API: USUARIOS ---

@app.route('/api/usuarios', methods=['GET'])
def get_usuarios():
    users = Usuario.query.all()
    resultado = [{"id": u.id, "usuario": u.usuario, "nombres": u.nombres, "rol": u.rol, "activo": u.activo, "tiene_password": bool(u.password_hash)} for u in users]
    return jsonify(resultado)

@app.route('/api/usuarios', methods=['POST'])
def add_usuario():
    data = request.json
    try:
        nuevo = Usuario(
            usuario=data['usuario'].strip(),
            nombres=data['nombres'].strip(),
            rol=data['rol'].strip(),
            activo=data.get('activo', True)
        )
        if data.get('password'):
            nuevo.password_hash = generate_password_hash(data['password'].strip())
        db.session.add(nuevo)
        db.session.commit()
        return jsonify({"success": True, "id": nuevo.id})
    except Exception as e:
        db.session.rollback()
        return jsonify({"success": False, "message": "Error al guardar (usuario debe ser único)."}), 400

@app.route('/api/usuarios/<int:id>', methods=['PUT', 'DELETE'])
def manage_usuario(id):
    user = db.session.get(Usuario, id)
    if not user:
        return jsonify({"success": False, "message": "Usuario no encontrado"}), 404
        
    if request.method == 'DELETE':
        db.session.delete(user)
        db.session.commit()
        return jsonify({"success": True, "message": "Usuario eliminado"})
        
    if request.method == 'PUT':
        data = request.json
        if 'nombres' in data: user.nombres = data['nombres'].strip()
        if 'rol' in data: user.rol = data['rol'].strip()
        if 'activo' in data: user.activo = data['activo']
        if data.get('password'): user.password_hash = generate_password_hash(data['password'].strip())
        
        try:
            db.session.commit()
            return jsonify({"success": True})
        except Exception as e:
            db.session.rollback()
            return jsonify({"success": False, "message": str(e)}), 400

@app.route('/api/usuarios/<int:id>/restricciones', methods=['GET'])
def get_restricciones(id):
    user = db.session.get(Usuario, id)
    if not user:
        return jsonify({"success": False, "message": "Usuario no encontrado"}), 404
    resultado = [{"id": r.id, "columna": r.columna, "valores": json.loads(r.valores_json)} for r in user.restricciones]
    return jsonify(resultado)

@app.route('/api/usuarios/<int:id>/restricciones', methods=['POST'])
def add_restriccion(id):
    user = db.session.get(Usuario, id)
    if not user:
        return jsonify({"success": False, "message": "Usuario no encontrado"}), 404
    data = request.json
    columna = data.get('columna', '').strip()
    valores = data.get('valores', [])
    if not columna or not valores:
        return jsonify({"success": False, "message": "Columna y valores son requeridos"}), 400
    # Si ya existe una restriccion para esa columna, añade los nuevos valores sin duplicados
    existing = next((r for r in user.restricciones if r.columna == columna), None)
    if existing:
        valores_actuales = json.loads(existing.valores_json)
        # Combinar listas y quitar duplicados
        nuevos_valores = list(set(valores_actuales + valores))
        existing.valores_json = json.dumps(nuevos_valores)
    else:
        db.session.add(RestriccionUsuario(usuario_id=id, columna=columna, valores_json=json.dumps(valores)))
    db.session.commit()
    return jsonify({"success": True})

@app.route('/api/usuarios/restricciones/<int:rid>', methods=['DELETE'])
def delete_restriccion(rid):
    r = db.session.get(RestriccionUsuario, rid)
    if not r:
        return jsonify({"success": False}), 404
    db.session.delete(r)
    db.session.commit()
    return jsonify({"success": True})

# --- API: TABLAS (FILTROS MAESTROS) ---

@app.route('/api/filtros', methods=['GET'])
def get_filtros():
    filtros = FiltroMaestro.query.all()
    resultado = {}
    for f in filtros:
        if f.entidad not in resultado:
            resultado[f.entidad] = []
        resultado[f.entidad].append({"id": f.id, "valor": f.valor})
    return jsonify(resultado)

@app.route('/api/filtros', methods=['POST'])
def add_filtro():
    data = request.json
    # Respetamos el casing del usuario para la entidad y el valor
    nuevo = FiltroMaestro(entidad=data['entidad'].strip(), valor=data['valor'].strip())
    try:
        db.session.add(nuevo)
        db.session.commit()
        return jsonify({"success": True, "id": nuevo.id})
    except:
        db.session.rollback()
        return jsonify({"success": False, "message": "Valor ya existe en esta entidad"}), 400

@app.route('/api/filtros/bulk', methods=['POST'])
def bulk_filtros():
    data = request.json # { entidad: '...', valores: [...] }
    entidad = data['entidad'].strip()
    count = 0
    for v in data['valores']:
        val = str(v).strip() # Quitamos el .upper() para respetar el formato
        if not FiltroMaestro.query.filter_by(entidad=entidad, valor=val).first():
            db.session.add(FiltroMaestro(entidad=entidad, valor=val))
            count += 1
    db.session.commit()
    return jsonify({"success": True, "added": count})

@app.route('/api/filtros/clear', methods=['POST'])
def clear_filtros():
    entidad = request.json.get('entidad')
    if entidad:
        FiltroMaestro.query.filter_by(entidad=entidad.strip()).delete()
    else:
        FiltroMaestro.query.delete()
    db.session.commit()
    return jsonify({"success": True})

@app.route('/api/filtros/<int:id>', methods=['DELETE'])
def delete_filtro(id):
    filtro = db.session.get(FiltroMaestro, id)
    if filtro:
        db.session.delete(filtro)
        db.session.commit()
        return jsonify({"success": True})
    return jsonify({"success": False, "message": "No encontrado"}), 404

@app.route('/api/filtros/entidad/rename', methods=['POST'])
def rename_entidad():
    data = request.json
    vieja = data['vieja'].strip()
    nueva = data['nueva'].strip()
    filtros = FiltroMaestro.query.filter_by(entidad=vieja).all()
    for f in filtros:
        f.entidad = nueva
    db.session.commit()
    return jsonify({"success": True})

@app.route('/api/filtros/entidad/delete', methods=['POST'])
def delete_entidad():
    data = request.json
    entidad = data['entidad'].strip()
    FiltroMaestro.query.filter_by(entidad=entidad).delete()
    db.session.commit()
    return jsonify({"success": True})

# --- API: IMPORTAR ---

@app.route('/api/import', methods=['POST'])
@admin_required
def process_import():
    temp_path = None
    try:
        file = request.files.get('file')
        source_type = request.form.get('source_type') or 'planobraCSV'
        
        # Obtener configuración persistente si no se envía en el form
        filter_active_val = request.form.get('filter_active')
        if filter_active_val is None:
            conf = ConfiguracionGlobal.query.get('import_filter_active')
            filter_active = conf.valor == 'true' if conf else False
        else:
            filter_active = filter_active_val == 'true'
            

        
        if not file: return jsonify({"error": "No hay archivo"}), 400
        
        filename = file.filename.lower()
        print(f"DEBUG: [Fase 0] Guardando archivo temporal: {filename}")
        
        # Guardamos en disco para poder hacer dos pasadas sin gastar RAM
        with tempfile.NamedTemporaryFile(delete=False, suffix=os.path.splitext(filename)[1]) as tmp:
            file.save(tmp.name)
            temp_path = tmp.name
        
        filesize = os.path.getsize(temp_path)
        print(f"DEBUG: [Fase 0] Archivo guardado en {temp_path}. Tamaño: {filesize / 1024:.2f} KB")

        def get_rows_iter(path):
            count = 0
            # Intentar detectar el tipo real si la extensión engaña
            is_csv = path.lower().endswith('.csv')
            try:
                if is_csv:
                    # Probamos si es CSV real
                    with open(path, 'r', encoding='utf-8', errors='ignore') as f:
                        line = f.readline()
                        if ';' not in line and ',' not in line and '\t' not in line:
                            is_csv = False # Probablemente es Excel mal nombrado
                
                if is_csv:
                    print(f"DEBUG: Procesando como CSV: {path}")
                    with open(path, 'r', encoding='utf-8', errors='ignore') as f:
                        reader = csv.DictReader(f)
                        for row in reader:
                            count += 1
                            if count % 1000 == 0: print(f"DEBUG: Leyendo fila CSV {count}...")
                            yield {str(k).upper().strip(): str(v).strip() for k, v in row.items() if k}
                else:
                    print(f"DEBUG: Procesando como XLSX/BIFF: {path}")
                    wb = load_workbook(filename=path, read_only=True, data_only=True)
                    sheet = wb.active
                    headers = []
                    for i, row in enumerate(sheet.iter_rows(values_only=True)):
                        if i == 0:
                            headers = [str(h).upper().strip() if h else f"COL_{idx}" for idx, h in enumerate(row)]
                            continue
                        if not any(row): continue
                        count += 1
                        if count % 1000 == 0: print(f"DEBUG: Leyendo fila Excel {count}...")
                        yield dict(zip(headers, [str(v).strip() if v is not None else "" for v in row]))
                    wb.close()
            except Exception as e:
                print(f"ERROR EN ITERATOR: {e}")
                raise

        def normalize_itp(itp):
            if itp is None: return ""
            itp = str(itp).strip()
            if itp.lower() in ["", "nan", "none"]: return ""
            if itp.endswith('.0'): itp = itp[:-2]
            return itp

        # Fase 1: Cache de IDs de forma eficiente
        print("DEBUG: [Fase 1] Cacheando base de datos (Lazy Mode)...")
        cache_obras_id = {} 
        try:
            # Intentamos usar json_extract si el SQLite lo soporta
            q = text("SELECT id, json_extract(data_json, '$.ITEMPLAN') FROM gestion_obras")
            res_proxy = db.session.execute(q)
            for row in res_proxy:
                oid, itp = row[0], row[1]
                if itp: cache_obras_id[normalize_itp(itp)] = oid
        except Exception as e:
            print(f"DEBUG: json_extract falló, usando modo manual: {e}")
            # Fallback seguro: Iteramos sobre el generador, no sobre .all()
            q = text("SELECT id, data_json FROM gestion_obras")
            res_proxy = db.session.execute(q)
            for row in res_proxy:
                oid, d_json = row[0], row[1]
                try:
                    # Buscamos el ITEMPLAN sin cargar todo el dict si es posible
                    # (aquí sí cargamos para ser exactos, pero uno a la vez)
                    itp = json.loads(d_json).get('ITEMPLAN')
                    if itp: cache_obras_id[normalize_itp(itp)] = oid
                except: continue
            
        print(f"DEBUG: Cache listo con {len(cache_obras_id)} registros.")
        gc.collect()

        # Filtros maestros
        filtros_dict = {}
        if filter_active:
            for f in FiltroMaestro.query.all():
                e, v = f.entidad.strip().upper(), f.valor.strip().upper()
                if e not in filtros_dict: filtros_dict[e] = []
                filtros_dict[e].append(v)

        manual_cols_names = [c.nombre.upper().strip() for c in ColumnaManual.query.all()]
        imported, updated, discarded = 0, 0, 0
        
        if source_type == 'detalleplanCSV':
            print("DEBUG: [Fase 2] Agregando Detalle Plan (Streaming)...")
            mo_sums, po_sets, vr_sets = {}, {}, {}
            for row in get_rows_iter(temp_path):
                itp = normalize_itp(row.get('ITEMPLAN', ''))
                if not itp: continue
                # Lógica de agregación igual a la anterior pero sin DataFrames
                is_mo = any(str(row.get(c, '')).upper().startswith('MO_') for c in ['PO', 'AREA'])
                if is_mo and 'VALORIZ MANO DE OBRA' in row:
                    try: 
                        raw_val = str(row['VALORIZ MANO DE OBRA']).replace(',', '').strip()
                        val = float(raw_val) if raw_val else 0.0
                    except: val = 0.0
                    mo_sums[itp] = mo_sums.get(itp, 0.0) + val
                
                is_mat = any(str(row.get(c, '')).upper().startswith('MAT_') for c in ['PO', 'AREA'])
                if is_mat:
                    po_sets.setdefault(itp, set()).add(row.get('PO', ''))
                    vr_sets.setdefault(itp, set()).add(row.get('VR', ''))

            itps_to_update = list(set(list(mo_sums.keys()) + list(po_sets.keys())))
            for i in range(0, len(itps_to_update), 500):
                batch = itps_to_update[i:i+500]
                ids = [cache_obras_id[itp] for itp in batch if itp in cache_obras_id]
                if not ids: continue
                obras = GestionObra.query.filter(GestionObra.id.in_(ids)).all()
                for obra in obras:
                    data = json.loads(obra.data_json)
                    itp_k = normalize_itp(data.get('ITEMPLAN', ''))
                    if itp_k in mo_sums: data['VALORIZ MANO DE OBRA'] = mo_sums[itp_k]
                    if itp_k in po_sets: data['PO'] = ', '.join(sorted(list(po_sets[itp_k] - {"", None})))
                    if itp_k in vr_sets: data['VR'] = ', '.join(sorted(list(vr_sets[itp_k] - {"", None})))
                    obra.data_json = safe_json_dumps(data)
                    updated += 1
                db.session.commit()
                db.session.expunge_all() # LIBERAR RAM
                gc.collect()
            if temp_path and os.path.exists(temp_path): os.remove(temp_path)
            return jsonify({"imported": 0, "updated": updated, "source": source_type})

        # Modo General Streaming (3 Pasos: Actualizar -> Filtrar -> Insertar)
        print("DEBUG: [Fase 2] Streaming General (Flujo Inteligente)...")
        objs_to_insert = []
        count = 0
        for row in get_rows_iter(temp_path):
            count += 1
            itp = normalize_itp(row.get('ITEMPLAN', ''))
            if not itp: discarded += 1; continue
            
            # --- PASO 1: ACTUALIZACIÓN DE EXISTENTES (Sin filtros restrictivos) ---
            if itp in cache_obras_id:
                obra = db.session.get(GestionObra, cache_obras_id[itp])
                if obra:
                    data = json.loads(obra.data_json)
                    if source_type == 'manual_update':
                        for k, v in row.items():
                            if k in manual_cols_names: data[k] = v
                    else:
                        for k in ['ESTADO PLAN', 'SUBESTADO TRUNCO', 'ESTADO', 'SITUACION']:
                            if k in row: data[k] = row[k]
                    obra.data_json = safe_json_dumps(data)
                    updated += 1
                
                # Procesar batch cada 500 filas de actualización para no saturar memoria
                if count % 500 == 0:
                    db.session.commit()
                    db.session.expunge_all()
                    gc.collect()
                continue # Pasar a la siguiente obra

            # --- PASO 2: FILTRADO DE NUEVOS ---
            if source_type == 'planobraCSV':
                if filter_active:
                    skip = False
                    for entidad, valores_permitidos in filtros_dict.items():
                        # Las llaves de row ya vienen en UPPER por get_rows_iter
                        val_fila = str(row.get(entidad, "")).strip().upper()
                        if val_fila and val_fila not in valores_permitidos:
                            # EL FILTRO DE SUBESTADO TRUNCO SOLO APLICA SI EL ESTADO PLAN ES 'TRUNCO'
                            if entidad == 'SUBESTADO TRUNCO':
                                estado_plan_val = str(row.get('ESTADO PLAN', '')).strip().upper()
                                if estado_plan_val != 'TRUNCO':
                                    continue # Pasar al siguiente filtro maestro sin descartar
                            
                            skip = True
                            break
                    if skip:
                        discarded += 1
                        continue
                
                # --- PASO 3: INSERCION DE NUEVOS ---
                objs_to_insert.append({"data_json": safe_json_dumps(row)})
                imported += 1
            else:
                discarded += 1

            if count % 500 == 0:
                if objs_to_insert:
                    db.session.execute(GestionObra.__table__.insert(), objs_to_insert)
                    objs_to_insert = []
                db.session.commit()
                db.session.expunge_all() # LIBERAR RAM
                gc.collect()
                print(f"DEBUG: Batch {count} procesado y guardado.")

        if objs_to_insert: db.session.execute(GestionObra.__table__.insert(), objs_to_insert)
        db.session.commit()
        return jsonify({"imported": imported, "updated": updated, "discarded": discarded, "source": source_type})

    except Exception as e:
        import traceback
        err = traceback.format_exc()
        print(f"ERROR CRITICO EN IMPORT: {e}\n{err}")
        return jsonify({"error": str(e), "trace": err}), 500
    finally:
        if temp_path and os.path.exists(temp_path):
            try:
                os.remove(temp_path)
                print(f"DEBUG: Archivo temporal eliminado: {temp_path}")
            except Exception as ex:
                print(f"AVISO: No se pudo eliminar temp: {ex}")


# --- API: COLUMNAS MANUALES ---

@app.route('/api/config/manual-columns', methods=['GET'])
def get_manual_columns():
    cols = ColumnaManual.query.all()
    return jsonify([{
        "id": c.id,
        "nombre": c.nombre,
        "tipo": c.tipo,
        "opciones": c.opciones
    } for c in cols])

@app.route('/api/config/manual-columns', methods=['POST'])
def add_manual_column():
    data = request.json
    nueva = ColumnaManual(
        nombre=data['nombre'].strip(),
        tipo=data['tipo'],
        opciones=data.get('opciones', '')
    )
    try:
        db.session.add(nueva)
        db.session.commit()
        return jsonify({"success": True, "id": nueva.id})
    except Exception as e:
        db.session.rollback()
        return jsonify({"success": False, "message": str(e)}), 400

@app.route('/api/config/manual-columns/<int:id>', methods=['PUT', 'DELETE'])
def manage_manual_column(id):
    col = db.session.get(ColumnaManual, id)
    if not col:
        return jsonify({"success": False, "message": "Columna no encontrada"}), 404
        
    if request.method == 'DELETE':
        db.session.delete(col)
        db.session.commit()
        return jsonify({"success": True})
        
    if request.method == 'PUT':
        data = request.json
        if 'nombre' in data: col.nombre = data['nombre'].strip()
        if 'tipo' in data: col.tipo = data['tipo']
        if 'opciones' in data: col.opciones = data.get('opciones', '')
        
        try:
            db.session.commit()
            return jsonify({"success": True})
        except Exception as e:
            db.session.rollback()
            return jsonify({"success": False, "message": str(e)}), 400

@app.route('/api/config/manual-columns/template')
@login_required
def download_manual_template():
    try:
        cols = ColumnaManual.query.all()
        header = ['ITEMPLAN'] + [c.nombre for c in cols]
        df = pd.DataFrame(columns=header)
        
        output = io.BytesIO()
        try:
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Plantilla_Manual')
        except:
            with pd.ExcelWriter(output) as writer:
                df.to_excel(writer, index=False, sheet_name='Plantilla_Manual')
                
        output.seek(0)
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='plantilla_columnas_manuales.xlsx'
        )
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500

# --- API: CONFIGURACIÓN DE FILTROS ---

@app.route('/api/config/filtros', methods=['GET'])
def get_config_filtros():
    configs = ConfiguracionFiltro.query.all()
    # Si no hay ninguna config, devolvemos un objeto vacío para no romper
    return jsonify({c.columna: {"tipo": c.tipo, "virtual_cols": json.loads(c.virtual_cols_json or '[]')} for c in configs})

@app.route('/api/config/filtros', methods=['POST'])
def update_config_filtro():
    data = request.json
    columna = data.get('columna')
    tipo = data.get('tipo')
    v_cols = data.get('virtual_cols')
    
    config = ConfiguracionFiltro.query.filter_by(columna=columna).first()
    if config:
        if tipo: config.tipo = tipo
        if v_cols is not None: config.virtual_cols_json = json.dumps(v_cols)
    else:
        config = ConfiguracionFiltro(
            columna=columna, 
            tipo=tipo if tipo else 'search',
            virtual_cols_json=json.dumps(v_cols) if v_cols is not None else '[]'
        )
        db.session.add(config)
    db.session.commit()
    return jsonify({"success": True})

# --- API: FILTROS VIRTUALES ---

@app.route('/api/config/filtros-virtuales', methods=['GET'])
def get_filtros_virtuales():
    filtros = FiltroVirtual.query.all()
    return jsonify([f.nombre for f in filtros])

@app.route('/api/config/filtros-virtuales', methods=['POST'])
def add_filtro_virtual():
    nombre = request.json.get('nombre', '').strip().upper()
    if not nombre: return jsonify({"success": False}), 400
    if not FiltroVirtual.query.filter_by(nombre=nombre).first():
        db.session.add(FiltroVirtual(nombre=nombre))
        db.session.commit()
    return jsonify({"success": True})

@app.route('/api/config/filtros-virtuales/<string:nombre>', methods=['DELETE'])
def delete_filtro_virtual(nombre):
    FiltroVirtual.query.filter_by(nombre=nombre).delete()
    db.session.commit()
    return jsonify({"success": True})

# --- API: MAPEOS DE FILTROS ---

@app.route('/api/config/mapeos', methods=['GET'])
def get_mapeos():
    mapeos = MapeoFiltro.query.all()
    return jsonify([{
        "id": m.id,
        "columna_criterio": m.columna_criterio,
        "valor_criterio": m.valor_criterio,
        "valores": json.loads(m.valores_json)
    } for m in mapeos])

@app.route('/api/config/mapeos', methods=['POST'])
def update_mapeo():
    data = request.json
    cc = data.get('columna_criterio')
    vc = data.get('valor_criterio')
    valores = data.get('valores', {})
    
    mapeo = MapeoFiltro.query.filter_by(columna_criterio=cc, valor_criterio=vc).first()
    if mapeo:
        mapeo.valores_json = json.dumps(valores)
    else:
        mapeo = MapeoFiltro(columna_criterio=cc, valor_criterio=vc, valores_json=json.dumps(valores))
        db.session.add(mapeo)
    
    db.session.commit()
    return jsonify({"success": True})

@app.route('/api/config/mapeos/<int:id>', methods=['DELETE'])
def delete_mapeo(id):
    MapeoFiltro.query.filter_by(id=id).delete()
    db.session.commit()
    return jsonify({"success": True})

# --- HELPERS PARA DATOS ---

def augment_virtual_columns(row_dict, mapeos, filter_configs):
    """
    Agrega columnas virtuales a un diccionario de datos basado en mapeos y configuraciones.
    """
    # Identificar todas las columnas virtuales posibles
    v_cols_set = new_set = set()
    for cfg in filter_configs.values():
        for vc in cfg.get('virtual_cols', []):
            v_cols_set.add(vc)
    
    for vc in v_cols_set:
        if vc not in row_dict:
            row_dict[vc] = ""
        
        # Recorremos todos los mapeos para esta columna virtual
        for m in mapeos:
            crit_col = m.columna_criterio
            crit_val = str(m.valor_criterio).strip()
            
            row_val = str(row_dict.get(crit_col, "")).strip()
            if row_val == crit_val:
                # Verificar si este criterio permite esta columna virtual según la config
                cfg = filter_configs.get(crit_col, {})
                allowed_cols = cfg.get('virtual_cols', [])
                
                mapeo_valores = json.loads(m.valores_json)
                if vc in allowed_cols and vc in mapeo_valores:
                    row_dict[vc] = mapeo_valores[vc]
    return row_dict

# --- API: ACTUALIZACIÓN DE PROYECTO ---

@app.route('/api/proyectos/update-field', methods=['POST'])
@login_required
def update_proyecto_field():
    user_rol = session.get('user_rol')
    if user_rol not in ['Admin', 'Administrador', 'Editor']:
        return jsonify({"success": False, "message": "No autorizado"}), 403

    data = request.json
    id_obra = data.get('id')
    field = data.get('field')
    value = data.get('value')
    
    obra = GestionObra.query.get(id_obra)
    if not obra:
        return jsonify({"success": False, "message": "Proyecto no encontrado"}), 404
        
    data_actual = json.loads(obra.data_json)
    data_actual[field] = value
    obra.data_json = json.dumps(data_actual)
    db.session.commit()
    return jsonify({"success": True})

# --- API: PROYECTOS ---

# --- API: CONFIGURACIÓN GLOBAL ---

@app.route('/api/config/global', methods=['GET'])
@admin_required
def get_global_config():
    configs = ConfiguracionGlobal.query.all()
    return jsonify({c.clave: c.valor for c in configs})

@app.route('/api/config/global', methods=['POST'])
@admin_required
def update_global_config():
    data = request.json
    for clave, valor in data.items():
        conf = ConfiguracionGlobal.query.get(clave)
        if conf:
            conf.valor = str(valor)
        else:
            db.session.add(ConfiguracionGlobal(clave=clave, valor=str(valor)))
    db.session.commit()
    return jsonify({"success": True})

# --- API: PROYECTOS ---

def get_filtered_projects(only_cerradas=False):
    user_id = session.get('user_id')
    user_rol = session.get('user_rol')
    user = db.session.get(Usuario, user_id)
    
    # Obtener restricciones del usuario (si no es Admin)
    restricciones = {}
    if user_rol != 'Admin' and user:
        for r in user.restricciones:
            restricciones[r.columna] = json.loads(r.valores_json)

    # Pre-cargar mapeos y configuraciones para columnas virtuales
    mapeos = MapeoFiltro.query.all()
    configs = {c.columna: {"tipo": c.tipo, "virtual_cols": json.loads(c.virtual_cols_json or '[]')} 
               for c in ConfiguracionFiltro.query.all()}

    # En lugar de usar el ORM que carga todo en RAM, usamos raw SQL y cursor para evitar MemoryError en Render
    q = text("SELECT id, data_json FROM gestion_obras")
    res_proxy = db.session.execute(q)
    
    resultado = []
    hoy = datetime.now()
    
    # Obtener Estados Permitidos (Filtro Maestro) para Separación Dinámica
    estados_permitidos = [str(f.valor).strip().upper() for f in FiltroMaestro.query.filter_by(entidad='ESTADO PLAN').all()]

    for row in res_proxy:
        try:
            oid = row[0]
            data = json.loads(row[1])
            data = augment_virtual_columns(data, mapeos, configs)
            
            # Separación Dinámica: Proyectos vs Cerradas
            estado_plan = str(data.get('ESTADO PLAN', '')).strip().upper()
            is_cerrada = False
            
            if estados_permitidos and (estado_plan not in estados_permitidos):
                is_cerrada = True
                
            if only_cerradas and not is_cerrada: continue
            if not only_cerradas and is_cerrada: continue

            # Aplicar filtrado por restricciones
            skip = False
            if restricciones:
                for col, valores_permitidos in restricciones.items():
                    val_obra = str(data.get(col, "")).strip()
                    if val_obra not in valores_permitidos:
                        skip = True
                        break
            if skip: continue


            data['__db_id'] = oid
            
            # Cálculo de Antigüedad (Días)
            antiguedad = ""
            fecha_creacion = data.get('FECHA CREACION IP')
            if fecha_creacion:
                try:
                    f_obj = None
                    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y"):
                        try:
                            f_obj = datetime.strptime(str(fecha_creacion).strip(), fmt)
                            break
                        except: continue
                    if f_obj:
                        diff = hoy - f_obj
                        antiguedad = max(0, diff.days)
                except: pass
            
            data['TIMING'] = antiguedad
            resultado.append(data)
            
            # Optimización de memoria en Render (Plan Free)
            if len(resultado) % 500 == 0:
                gc.collect()
        except Exception as e:
            print(f"Error procesando obra {oid}: {e}")
            continue
    return resultado

@app.route('/api/proyectos', methods=['GET'])
@login_required
def get_proyectos():
    return jsonify(get_filtered_projects(only_cerradas=False))

@app.route('/api/cerradas', methods=['GET'])
@login_required
def get_cerradas():
    return jsonify(get_filtered_projects(only_cerradas=True))

@app.route('/api/proyectos/clear', methods=['POST'])
def clear_proyectos():
    GestionObra.query.delete()
    db.session.commit()
    return jsonify({"success": True})

@app.route('/api/proyectos/export', methods=['POST'])
def export_proyectos():
    data = request.json # Data filtrada desde el cliente
    if not data: return jsonify({"error": "No hay datos para exportar"}), 400
    
    df = pd.DataFrame(data)
    if '__db_id' in df.columns: df = df.drop(columns=['__db_id'])
    
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False, sheet_name='Consolidado')
    
    output.seek(0)
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', 
                     as_attachment=True, download_name=f'Consolidado_ANTON_{datetime.now().strftime("%Y%m%d")}.xlsx')

if __name__ == '__main__':
    port = int(os.environ.get("PORT", 5000))
    app.run(host='0.0.0.0', port=port, debug=False)
